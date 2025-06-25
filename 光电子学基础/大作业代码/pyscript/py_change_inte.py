# 读取xlsx修改采样间隔
import openpyxl
import os

def process_excel(file_name, n):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_name)
    new_wb = openpyxl.Workbook()  # 创建一个新的工作簿
    new_wb.remove(new_wb.active)  # 删除默认的空表

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        new_sheet = new_wb.create_sheet(sheet_name)  # 在新工作簿中创建对应的表

        # 读取表格内容
        wavelengths = []
        responses = []

        for row in sheet.iter_rows(min_row=2, max_row=702, min_col=1, max_col=2, values_only=True):
            wavelengths.append(row[0])
            responses.append(row[1])

        # 采样处理
        sampled_wavelengths = wavelengths[::n]
        sampled_responses = responses[::n]

        # 写入新表格
        new_sheet.append([sheet.cell(1, 1).value, sheet.cell(1, 2).value])  # 写入表头
        for wl, resp in zip(sampled_wavelengths, sampled_responses):
            new_sheet.append([wl, resp])

    # 保存修改后的文件
    output_file_name = f"processed_{file_name}"
    new_wb.save(output_file_name)
    print(f"处理完成，保存为 {output_file_name}")

if __name__ == "__main__":
    # 获取当前目录下的 xlsx 文件
    current_dir = os.getcwd()
    xlsx_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]

    if not xlsx_files:
        print("当前目录下没有找到 xlsx 文件！")
    else:
        print("找到以下 xlsx 文件：")
        for i, file in enumerate(xlsx_files):
            print(f"{i + 1}. {file}")

        # 选择文件
        file_index = int(input("请输入要处理的文件编号：")) - 1
        selected_file = xlsx_files[file_index]

        # 输入采样间隔 n
        n = int(input("请输入采样间隔 n："))

        # 处理文件
        process_excel(selected_file, n)